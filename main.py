from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
import os


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with your domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class Submission(BaseModel):
    product: str
    quantity: float
    farmer_name: str
    phone: str
    bank_name: str
    account_no: str
    ifsc: str
    price_per_kg: float

EXCEL_FILE = "submissions.xlsx"

@app.on_event("startup")
def init_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Product", "Quantity", "Farmer Name", "Phone", "Bank Name", "Account No", "IFSC", "Price/kg", "Amount"])
        wb.save(EXCEL_FILE)

@app.post("/submit")
def submit(data: Submission):
    net_qty = round(data.quantity * 0.95, 2)
    amount = round(net_qty * data.price_per_kg, 2)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.product, data.quantity, data.farmer_name, data.phone,
        data.bank_name, data.account_no, data.ifsc,
        data.price_per_kg, amount
    ])
    wb.save(EXCEL_FILE)

    return {"message": "Submitted", "net_quantity": net_qty, "amount": amount}
